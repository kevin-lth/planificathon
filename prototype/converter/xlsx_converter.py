from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, colors, Font, Color, fills
from openpyxl.utils.cell import range_boundaries
import json
from dataclasses import dataclass
from dataclasses_json import dataclass_json

from django.core.files.temp import NamedTemporaryFile

@dataclass_json
@dataclass
class Journee:
    besoin_matin: int
    besoin_soir: int
    besoin_nuit: int
    besoin_sve: int
    nb_jca: int
    matin: list
    soir: list
    nuit: list
    sve: list
    jca: list

centrer = Alignment(horizontal='center', vertical='center')

def dict_to_xlsx(data, then):
    liste_des_agents = data["agents"]
    dictjours = data["planning"]

    nb_agents = len(liste_des_agents)
    nb_jours = len(dictjours)

    wk = Workbook()
    del wk['Sheet']
    sheet = wk.create_sheet(title='Feuil1')
    redft = Font(color="00D00000", bold=True)
    monjolicyan = colors.Color(rgb='0099CCFF')
    remplircyan = fills.PatternFill(patternType='solid', fgColor=monjolicyan)
    monjoliturquoise = colors.Color(rgb='0000FFFF')
    remplirturquoise = fills.PatternFill(patternType='solid', fgColor=monjoliturquoise)
    monjolijaune = colors.Color(rgb='00FFFF99')
    remplirjaune = fills.PatternFill(patternType='solid', fgColor=monjolijaune)

    sheet["A2"] = "Agent"
    sheet["A2"].alignment = centrer
    sheet["A2"].fill = remplirturquoise
    sheet["B2"] = "%"
    sheet["B2"].alignment = centrer
    sheet["B2"].fill = remplirturquoise
    sheet["C2"] = "J/N"
    sheet["C2"].alignment = centrer
    sheet["C2"].fill = remplirturquoise

    for r in range(3, nb_agents + 3):
        cellule = sheet.cell(row=r, column=1)
        cellule.value = r-2
        cellule.alignment = centrer
        cellule.fill = remplirjaune
    for r in range(3, nb_agents + 3):
        cellule = sheet.cell(row=r, column=2)
        cellule.value = liste_des_agents[r-3]['pourcentage']
        cellule.alignment = centrer
        cellule.fill = remplirjaune
    for r in range(3, nb_agents + 3):
        cellule = sheet.cell(row=r, column=3)
        cellule.value = 'J'
        cellule.alignment = centrer
        cellule.fill = remplirjaune

    besoinstext = ['Besoins Matin', 'Besoins Soir', 'Besoins Nuit', 'Besoins Sve', 'Nb Jca']
    for r in range(3 + nb_agents, nb_agents + 3 + 5):
        cellule = sheet.cell(row=r, column=1)
        cellule.value = besoinstext[r - 3 - nb_agents]
        cellule.alignment = centrer
        sheet.merge_cells(start_row = r, end_row=r, start_column=1, end_column=3)

    for week in range(nb_jours//7):
        firstc = week*7 + 4
        lastc = (week + 1)*7 + 4 -1
        sheet.merge_cells(start_row=1, end_row=1, start_column=firstc, end_column=lastc)
        cellule = sheet.cell(column=firstc, row=1)
        cellule.value = f"Semaine {week + 1}"
        cellule.alignment = centrer

    jours_de_la_semaine = ['L', 'M', 'M', 'J', 'V', 'S', 'D']
    for day in range(nb_jours):
        numc = day + 4
        cellule = sheet.cell(column=numc, row=2)
        cellule.value = jours_de_la_semaine[day%7]
        cellule.font = redft
        cellule.fill = remplircyan
        cellule.alignment = centrer

    dict_convert_m = {'matin':'M', 'soir':'S', 'nuit':'N', 'jca':'Jca','sve':'Sve'}
    list_convert = ['besoin_matin', 'besoin_soir', 'besoin_nuit', 'besoin_sve', 'nb_jca']
    for day in range(nb_jours):
        firstr = 3
        numc = day + 4
        for r,a in enumerate(liste_des_agents):
            cellule = sheet.cell(row=firstr + r, column=numc)
            cellule.value = '.'
            for key_word in dict_convert_m.keys():
                if a[key_word][day]:
                    cellule.value = dict_convert_m[key_word]
            cellule.alignment = centrer
        for r,besoin in enumerate(list_convert):
            cellule = sheet.cell(row=firstr + nb_agents + r, column=numc)
            cellule.value = dictjours[day][besoin]
            cellule.alignment = centrer

    with NamedTemporaryFile(delete=True, suffix=".xlsx") as tmp:
        wk.save(tmp.name)
        result = then(tmp.name)
        return result

def xlsx_to_dict(filename, sheetname):

    wk = load_workbook(filename=filename, data_only=True)

    sheet  = wk[sheetname]

    row_count = sheet.max_row
    column_count = sheet.max_column

    liste_des_agents = []
    nb_agents = 0
    agent_row_start = None
    agent_column_start = None
    for i in range(1, row_count):
        for j in range(1, column_count):
            if agent_column_start and agent_column_start < j:
                break
            if sheet.cell(row=i, column=j).value:
                if isinstance(sheet.cell(row=i, column=j).value, int):
                    if not agent_row_start:
                        agent_row_start = i
                    if not agent_column_start:
                        agent_column_start = j
                    nb_agents += 1
                    liste_des_agents.append(dict({'numero': nb_agents, 'pourcentage':sheet.cell(row=i,column=j+1).value}))
                    break
    for r in range(agent_row_start, agent_row_start+nb_agents):
        matin = [1 if (sheet.cell(column=c, row=r).value == 'M') else 0 for c in range(4, column_count+1)]
        soir = [1 if (sheet.cell(column=c, row=r).value == 'S') else 0 for c in range(4, column_count+1)]
        nuit = [1 if (sheet.cell(column=c, row=r).value == 'N') else 0 for c in range(4, column_count+1)]
        sve = [1 if (sheet.cell(column=c, row=r).value == 'Sve') else 0 for c in range(4, column_count+1)]
        jca = [1 if (sheet.cell(column=c, row=r).value == 'Jca') else 0 for c in range(4, column_count+1)]
        i = r - agent_row_start
        liste_des_agents[i]['matin'] = matin
        liste_des_agents[i]['soir'] = soir
        liste_des_agents[i]['nuit'] = nuit
        liste_des_agents[i]['sve'] = sve
        liste_des_agents[i]['jca'] = jca

    besoins_matin = [sheet.cell(column=c, row=nb_agents+agent_row_start).value for c in range(4, column_count+1)]
    besoins_soir = [sheet.cell(column=c, row=nb_agents+agent_row_start+1).value for c in range(4, column_count+1)]
    besoins_nuit = [sheet.cell(column=c, row=nb_agents+agent_row_start+2).value for c in range(4, column_count+1)]
    besoins_sve = [sheet.cell(column=c, row=nb_agents+agent_row_start+3).value for c in range(4, column_count+1)]
    nb_jca = [sheet.cell(column=c, row=nb_agents+agent_row_start+4).value for c in range(4, column_count+1)]
    for k in range(len(nb_jca)):
        if nb_jca[k] is None:
            nb_jca[k] = 0
    planning = []
    for d in range(column_count - 3):
        bm = besoins_matin[d]
        bs = besoins_soir[d]
        bn = besoins_nuit[d]
        bv = besoins_sve[d]
        nj = nb_jca[d]
        matin_ass = []
        soir_ass = []
        nuit_ass = []
        sve_ass = []
        jca_ass = []
        for a in liste_des_agents:
            if a['matin'][d]:
                matin_ass.append(a['numero'])
            if a['soir'][d]:
                soir_ass.append(a['numero'])
            if a['nuit'][d]:
                nuit_ass.append(a['numero'])
            if a['sve'][d]:
                sve_ass.append(a['numero'])
            if a['jca'][d]:
                jca_ass.append(a['numero'])
        jour = Journee(bm,bs,bn,bv,nj,matin_ass,soir_ass,nuit_ass,sve_ass,jca_ass)
        planning.append(jour)

    lsjours =[]
    for j in planning:
        lsjours.append(j.to_dict())
    return { "planning": lsjours, "agents": liste_des_agents }
