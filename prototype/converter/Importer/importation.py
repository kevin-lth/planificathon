from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
import json
from dataclasses import dataclass
from dataclasses_json import dataclass_json


wk = load_workbook(filename="./ExemplesCHU.xlsx", data_only=True)

sheet  = wk['Feuil1']

row_count = sheet.max_row
column_count = sheet.max_column

# for cell_group in sheet.merged_cells.ranges:
#     min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
#     print(str(cell_group))
#     print(min_col, max_col)
#     top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
#     sheet.unmerge_cells(str(cell_group))
#     for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
#         for cell in row:
#             cell.value = top_left_cell_value

# for r in sheet.iter_rows(values_only=True):
#     print(r)
# for i in range(4):
#     for j in range(10):
#         print(sheet.cell(i+1,j+1).value)

# def trouver_mots_clefs(sheet):
#     semaine_mots = ['semaine']
#     agent_mots = ['agent','nom']
#     besoin_matin_mots = ['besoins m']
#     besoin_soir_mots = ['besoins s']
#     besoin_nuit_mots = ['besoins n']
#     besoin_sve_mots = ['besoins sve']
#     nb_jca_mots = ['nb jca']
@dataclass_json
@dataclass
class Journee:
    besoin_matin: int
    besoin_soir: int
    besoin_nuit: int
    besoin_sve: int
    nb_jca: int
    matin: list
    soir: list
    nuit: list
    sve: list
    jca: list

# @dataclass
# class Agent:
#     numero : int
#     pourcentage: int
#     matins: list
#     soirs: list
#     nuit: list
#     sve: list
#     jca: list



liste_des_agents = []
nb_agents = 0
agent_row_start = None
for i in range(1, row_count):
    if sheet.cell(row=i, column=1).value:
        if isinstance(sheet.cell(row=i, column=1).value, int):
            if not agent_row_start:
                agent_row_start = i
            nb_agents += 1
            liste_des_agents.append(dict({'numero': nb_agents, 'pourcentage':sheet.cell(row=i,column=2).value}))

for r in range(agent_row_start, agent_row_start+nb_agents):
    matin = [1 if (sheet.cell(column=c, row=r).value == 'M') else 0 for c in range(4, column_count)]
    soir = [1 if (sheet.cell(column=c, row=r).value == 'S') else 0 for c in range(4, column_count)]
    nuit = [1 if (sheet.cell(column=c, row=r).value == 'N') else 0 for c in range(4, column_count)]
    sve = [1 if (sheet.cell(column=c, row=r).value == 'Sve') else 0 for c in range(4, column_count)]
    jca = [1 if (sheet.cell(column=c, row=r).value == 'Jca') else 0 for c in range(4, column_count)]

    i = r - agent_row_start
    liste_des_agents[i]['matin'] = matin
    liste_des_agents[i]['soir'] = soir
    liste_des_agents[i]['nuit'] = nuit
    liste_des_agents[i]['sve'] = sve
    liste_des_agents[i]['jca'] = jca

besoins_matin = [sheet.cell(column=c, row=nb_agents+agent_row_start).value for c in range(4, column_count)]
besoins_soir = [sheet.cell(column=c, row=nb_agents+agent_row_start+1).value for c in range(4, column_count)]
besoins_nuit = [sheet.cell(column=c, row=nb_agents+agent_row_start+2).value for c in range(4, column_count)]
besoins_sve = [sheet.cell(column=c, row=nb_agents+agent_row_start+3).value for c in range(4, column_count)]
nb_jca = [sheet.cell(column=c, row=nb_agents+agent_row_start+4).value for c in range(4, column_count)]
print(nb_jca)
for k in range(len(nb_jca)):
    if nb_jca[k] is None:
        nb_jca[k] = 0


planning = []
for d in range(column_count - 4):
    bm = besoins_matin[d]
    bs = besoins_soir[d]
    bn = besoins_nuit[d]
    bv = besoins_sve[d]
    nj = nb_jca[d]
    matin_ass = []
    soir_ass = []
    nuit_ass = []
    sve_ass = []
    jca_ass = []
    for a in liste_des_agents:
        if a['matin'][d]:
            matin_ass.append(a['numero'])
        if a['soir'][d]:
            soir_ass.append(a['numero'])
        if a['nuit'][d]:
            nuit_ass.append(a['numero'])
        if a['sve'][d]:
            sve_ass.append(a['numero'])
        if a['jca'][d]:
            jca_ass.append(a['numero'])


    jour = Journee(bm,bs,bn,bv,nj,matin_ass,soir_ass,nuit_ass,sve_ass,jca_ass)
    planning.append(jour)



# print(nb_agents)
# print(liste_des_agents)
with open('planning.json', 'w') as jsfile:
    lsjours =[]
    for j in planning:
        lsjours.append(j.to_dict())
    json.dump(lsjours, jsfile)
with open('agents.json', 'w') as jsfile:
    json.dump(liste_des_agents, jsfile)
