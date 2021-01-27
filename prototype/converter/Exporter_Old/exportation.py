from openpyxl import Workbook
from openpyxl.styles import Alignment, colors, Font, Color, fills
from openpyxl.utils.cell import range_boundaries
import json
from dataclasses import dataclass
from dataclasses_json import dataclass_json

@dataclass_json
@dataclass
class Journee:
    besoin_matin: int
    besoin_soir: int
    besoin_nuit: int
    besoin_sve: int
    nb_jca: int
    matin: list
    soir: list
    nuit: list
    sve: list
    jca: list

centrer = Alignment(horizontal='center', vertical='center')

def ecrire_xlxs(path_agentjs, path_planjs, outputxlxs):

    with open(path_agentjs, 'r') as jsfile:
        liste_des_agents = json.load(jsfile)
    with open(path_planjs, 'r') as jsfile:
        dictjours = json.load(jsfile)

    nb_agents = len(liste_des_agents)
    nb_jours = len(dictjours)

    wk = Workbook()
    del wk['Sheet']
    sheet = wk.create_sheet(title='Feuil1')
    redft = Font(color="00D00000", bold=True)
    monjolicyan = colors.Color(rgb='0099CCFF')
    remplircyan = fills.PatternFill(patternType='solid', fgColor=monjolicyan)
    monjoliturquoise = colors.Color(rgb='0000FFFF')
    remplirturquoise = fills.PatternFill(patternType='solid', fgColor=monjoliturquoise)
    monjolijaune = colors.Color(rgb='00FFFF99')
    remplirjaune = fills.PatternFill(patternType='solid', fgColor=monjolijaune)

    sheet["A2"] = "Agent"
    sheet["A2"].alignment = centrer
    sheet["A2"].fill = remplirturquoise
    sheet["B2"] = "%"
    sheet["B2"].alignment = centrer
    sheet["B2"].fill = remplirturquoise
    sheet["C2"] = "J/N"
    sheet["C2"].alignment = centrer
    sheet["C2"].fill = remplirturquoise

    for r in range(3, nb_agents + 3):
        cellule = sheet.cell(row=r, column=1)
        cellule.value = r-2
        cellule.alignment = centrer
        cellule.fill = remplirjaune
    for r in range(3, nb_agents + 3):
        cellule = sheet.cell(row=r, column=2)
        cellule.value = liste_des_agents[r-3]['pourcentage']
        cellule.alignment = centrer
        cellule.fill = remplirjaune
    for r in range(3, nb_agents + 3):
        cellule = sheet.cell(row=r, column=3)
        cellule.value = 'J'
        cellule.alignment = centrer
        cellule.fill = remplirjaune

    besoinstext = ['Besoins Matin', 'Besoins Soir', 'Besoins Nuit', 'Besoins Sve', 'Nb Jca']
    for r in range(3 + nb_agents, nb_agents + 3 + 5):
        cellule = sheet.cell(row=r, column=1)
        cellule.value = besoinstext[r - 3 - nb_agents]
        cellule.alignment = centrer
        sheet.merge_cells(start_row = r, end_row=r, start_column=1, end_column=3)

    for week in range(nb_jours//7):
        firstc = week*7 + 4
        lastc = (week + 1)*7 + 4 -1
        sheet.merge_cells(start_row=1, end_row=1, start_column=firstc, end_column=lastc)
        cellule = sheet.cell(column=firstc, row=1)
        cellule.value = f"Semaine {week + 1}"
        cellule.alignment = centrer

    jours_de_la_semaine = ['L', 'M', 'M', 'J', 'V', 'S', 'D']
    for day in range(nb_jours):
        numc = day + 4
        cellule = sheet.cell(column=numc, row=2)
        cellule.value = jours_de_la_semaine[day%7]
        cellule.font = redft
        cellule.fill = remplircyan
        cellule.alignment = centrer

    dict_convert_m = {'matin':'M', 'soir':'S', 'nuit':'N', 'jca':'Jca','sve':'Sve'}
    list_convert = ['besoin_matin', 'besoin_soir', 'besoin_nuit', 'besoin_sve', 'nb_jca']
    for day in range(nb_jours):
        firstr = 3
        numc = day + 4
        for r,a in enumerate(liste_des_agents):
            cellule = sheet.cell(row=firstr + r, column=numc)
            cellule.value = '.'
            for key_word in dict_convert_m.keys():
                if a[key_word][day]:
                    cellule.value = dict_convert_m[key_word]
            cellule.alignment = centrer
        for r,besoin in enumerate(list_convert):
            cellule = sheet.cell(row=firstr + nb_agents + r, column=numc)
            cellule.value = dictjours[day][besoin]
            cellule.alignment = centrer

    wk.save(filename=outputxlxs)


if __name__ == "__main__":
    chemin_plan = "../Importer/planning.json"
    chemin_agents = "../Importer/agents.json"
    filename = "super_planning.xlxs"
    ecrire_xlxs(chemin_agents, chemin_plan, filename)


